# -*- coding: utf-8 -*-
"""
Export Excel des rapports (section 5.3) — vérifie que `format=xlsx` renvoie un vrai classeur
téléchargeable, ouvrable par openpyxl, avec la feuille « Résumé » et une feuille dédiée pour
les listes (échéances). Complète le bouton « Exporter en Excel » ajouté à l'écran Reporting.

S'exécute APRÈS la recette (ordre alphabétique des fichiers) : la base contient alors la
police et la quittance du scénario de référence.
"""

import datetime
from io import BytesIO

from openpyxl import load_workbook

AUJ = datetime.date.today()
XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_export_reporting_xlsx_ouvrable(client, auth):
    debut = (AUJ - datetime.timedelta(days=1)).isoformat()
    fin = (AUJ + datetime.timedelta(days=1)).isoformat()

    # 1. Chiffre d'affaires en xlsx : classeur ouvrable, feuille « Résumé », en-tête et données
    r = client.get(f"/reporting/chiffre-affaires?date_debut={debut}&date_fin={fin}&format=xlsx",
                   headers=auth)
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == XLSX
    assert "attachment" in r.headers.get("content-disposition", "")
    wb = load_workbook(BytesIO(r.content))
    assert "Résumé" in wb.sheetnames, wb.sheetnames
    resume = wb["Résumé"]
    libelles = [resume.cell(row=i, column=1).value for i in range(1, resume.max_row + 1)]
    assert any(l and "quittances" in str(l).lower() for l in libelles), libelles

    # 2. Échéances en xlsx : une feuille dédiée pour la liste des polices (horizon large pour
    #    inclure la police de la recette, à échéance dans ~1 an).
    r = client.get("/reporting/echeances?jours=400&format=xlsx", headers=auth)
    assert r.status_code == 200, r.text
    wb = load_workbook(BytesIO(r.content))
    assert "Résumé" in wb.sheetnames, wb.sheetnames
    # la clé de liste « polices » devient une feuille dédiée
    assert any(nom.lower().startswith("police") for nom in wb.sheetnames), wb.sheetnames
