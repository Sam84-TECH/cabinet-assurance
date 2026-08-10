# -*- coding: utf-8 -*-
"""
Recette bout en bout du workflow métier décrit dans CLAUDE.md :
SOUS -> POL (quittance) -> ENC -> BANQ -> REV, puis contrôle du tableau de bord.

Un seul test déroule tout le scénario, via l'API HTTP réelle, sur la base de test isolée
(voir conftest.py). C'est le « cahier de recette » exigé au §9 du CDCF : quand il passe au
vert sans intervention manuelle, le backend est considéré terminé pour le périmètre auto.
"""

import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

AUJ = datetime.date.today()
DANS_UN_AN = AUJ + datetime.timedelta(days=365)


def dec(valeur):
    """Normalise une valeur JSON (nombre ou chaîne) en Decimal pour des comparaisons exactes."""
    return Decimal(str(valeur))


def test_recette_bout_en_bout(client, auth, refs):
    # 0. RF-CRM (§25) : un client sans identité obligatoire est refusé (422), pas accepté
    reponse = client.post("/clients", headers=auth, json={"type": "particulier"})
    assert reponse.status_code == 422, reponse.text
    reponse = client.post("/clients", headers=auth, json={"type": "entreprise"})  # sans raison sociale/ICE
    assert reponse.status_code == 422, reponse.text

    # 1. Client entreprise
    reponse = client.post("/clients", headers=auth, json={
        "type": "entreprise", "raison_sociale": "Recette Auto SARL", "ice": "RECETTE0001",
        "telephone": "0522000000", "ville": "Casablanca",
    })
    assert reponse.status_code == 200, reponse.text
    client_id = reponse.json()["id"]

    # 2. Police auto (produit paramétré par le seed)
    reponse = client.post("/sous/polices", headers=auth, json={
        "client_id": client_id, "produit_id": refs["produit_id"],
        "date_effet": AUJ.isoformat(), "date_echeance": DANS_UN_AN.isoformat(),
        "statut": "en_vigueur",  # tentative de contournement : ignorée côté serveur
    })
    assert reponse.status_code == 200, reponse.text
    police = reponse.json()
    police_id = police["id"]
    assert police["numero_police"].startswith("POL-")
    assert police["statut"] == "en_attente_effet", police  # statut forcé serveur

    # 3. Véhicule (risque) rattaché à la police
    reponse = client.post("/risques", headers=auth, json={
        "police_id": police_id, "type_risque": "vehicule",
        "attributs": {
            "immatriculation": "12345-A-6", "marque": "Dacia", "modele": "Logan",
            "date_mise_en_circulation": "03/2021", "valeur_neuf": "150000.00",
            "puissance_fiscale": 8,  # pour la tarification RC par barème de puissance (§24)
        },
    })
    assert reponse.status_code == 200, reponse.text
    risque_id = reponse.json()["id"]

    # 4. Garanties avec leurs primes : RC 800 + DC 200 = 1000 de prime nette
    for code, prime in (("RC", "800.00"), ("DC", "200.00")):
        reponse = client.post("/police-garanties", headers=auth, json={
            "police_id": police_id, "risque_id": risque_id,
            "garantie_id": refs["garanties"][code], "montant_prime": prime,
        })
        assert reponse.status_code == 200, reponse.text

    # 5. Avenant « affaire nouvelle »
    reponse = client.post("/sous/avenants", headers=auth, json={
        "police_id": police_id, "type_avenant": "affaire_nouvelle", "date_effet": AUJ.isoformat(),
        "statut": "valide", "valide_par": 999,  # tentative de contournement : ignorée côté serveur
    })
    assert reponse.status_code == 200, reponse.text
    assert reponse.json()["statut"] == "brouillon", reponse.json()  # statut forcé serveur
    avenant_id = reponse.json()["id"]

    # 5b. RF-SOUS-04 : l'émission est bloquée tant que les pièces obligatoires manquent
    reponse = client.patch(f"/sous/avenants/{avenant_id}/valider", headers=auth)
    assert reponse.status_code == 400, reponse.text
    assert "pi" in reponse.json()["detail"].lower()  # « pièces … manquantes »

    # 6. Fourniture des pièces justificatives obligatoires
    for piece_id in refs["pieces_oblig"]:
        reponse = client.post("/pieces-fournies", headers=auth, json={
            "police_id": police_id, "piece_requise_id": piece_id, "reference": "fournie",
        })
        assert reponse.status_code == 200, reponse.text

    # 7. Validation de l'avenant -> quittance générée automatiquement (RF-SOUS-07)
    reponse = client.patch(f"/sous/avenants/{avenant_id}/valider", headers=auth)
    assert reponse.status_code == 200, reponse.text
    validation = reponse.json()
    assert validation["statut"] == "valide"
    # §26 : la réponse de validation inclut la quittance générée (plus de GET séparé)
    assert validation["quittance"] is not None, validation
    assert dec(validation["quittance"]["prime_ttc"]) == Decimal("1140.00")
    # §27 : reste dû = TTC (rien réglé encore) sur la quittance renvoyée
    assert dec(validation["quittance"]["reste_du"]) == Decimal("1140.00")

    # 8. Quittance auto : prime_ttc = prime_nette + taxes + timbres + accessoires (règle n°4)
    reponse = client.get("/quittances", headers=auth, params={"police_id": police_id})
    assert reponse.status_code == 200, reponse.text
    quittances = reponse.json()
    assert len(quittances) == 1, quittances
    quittance = quittances[0]
    quittance_id = quittance["id"]
    assert dec(quittance["prime_nette"]) == Decimal("1000.00")
    assert dec(quittance["taxes"]) == Decimal("140.00")  # 14 % de la prime nette (défaut)
    assert dec(quittance["prime_ttc"]) == (
        dec(quittance["prime_nette"]) + dec(quittance["taxes"])
        + dec(quittance["timbres"]) + dec(quittance["accessoires"])
    )
    assert dec(quittance["prime_ttc"]) == Decimal("1140.00")
    assert dec(quittance["commission"]) == Decimal("100.00")  # 10 % (barème Sanlam/Auto)
    assert quittance["statut"] == "emise"
    # §27 : montant réglé et reste dû réels exposés (ici impayée : 0 réglé, TTC restant dû)
    assert dec(quittance["montant_regle"]) == Decimal("0.00")
    assert dec(quittance["reste_du"]) == Decimal("1140.00")

    # 8b. Tarification automatique par garantie (RF-SOUS-02) : montant_prime omis -> calculé au
    # barème du produit. Ajouté APRÈS la quittance pour ne pas modifier ses montants.
    reponse = client.post("/police-garanties", headers=auth, json={  # mode taux : 1,5 % de 150 000
        "police_id": police_id, "risque_id": risque_id,
        "garantie_id": refs["garanties"]["VOL"], "capital_assure": "150000.00",
    })
    assert reponse.status_code == 200, reponse.text
    assert dec(reponse.json()["montant_prime"]) == Decimal("2250.00"), reponse.json()

    reponse = client.post("/police-garanties", headers=auth, json={  # mode forfait : 220,00
        "police_id": police_id, "risque_id": risque_id, "garantie_id": refs["garanties"]["BDG"],
    })
    assert reponse.status_code == 200, reponse.text
    assert dec(reponse.json()["montant_prime"]) == Decimal("220.00"), reponse.json()

    # mode bareme_puissance (RC auto, §24) : la prime vient de la tranche couvrant la puissance
    # fiscale du véhicule (8 CV -> tranche <= 10 -> 2 400,00), sans capital assuré.
    reponse = client.post("/referentiel/garanties", headers=auth, json={
        "produit_id": refs["produit_id"], "code": "RC_TEST", "nom": "RC barème (test)",
        "parametres": {"mode": "bareme_puissance", "tranches": [
            {"puissance_max": 4, "montant": "1200.00"},
            {"puissance_max": 10, "montant": "2400.00"},
            {"puissance_max": None, "montant": "3200.00"},
        ]},
    })
    assert reponse.status_code == 200, reponse.text
    garantie_bareme_id = reponse.json()["id"]
    reponse = client.post("/police-garanties", headers=auth, json={
        "police_id": police_id, "risque_id": risque_id, "garantie_id": garantie_bareme_id,
    })
    assert reponse.status_code == 200, reponse.text
    assert dec(reponse.json()["montant_prime"]) == Decimal("2400.00"), reponse.json()

    # 9. Encaissement par chèque du montant TTC
    reponse = client.post("/encaissements", headers=auth, json={
        "client_id": client_id, "mode_paiement": "cheque", "montant": "1140.00",
        "date_encaissement": AUJ.isoformat(),
        "cheque_banque": "Attijariwafa Bank", "cheque_numero": "REC-0001",
        "statut": "rapproche_banque",  # tentative de contournement : ignorée côté serveur
    })
    assert reponse.status_code == 200, reponse.text
    assert reponse.json()["statut"] == "enregistre", reponse.json()  # statut forcé serveur
    encaissement_id = reponse.json()["id"]

    # 10. Affectation de l'encaissement à la quittance
    reponse = client.post(f"/encaissements/{encaissement_id}/affecter/{quittance_id}",
                          headers=auth, json={"montant_affecte": "1140.00"})
    assert reponse.status_code == 200, reponse.text

    # 11. La quittance passe en « réglée » ; reste dû à 0 (§27)
    reponse = client.get(f"/quittances/{quittance_id}", headers=auth)
    corps = reponse.json()
    assert corps["statut"] == "reglee", corps
    assert dec(corps["montant_regle"]) == Decimal("1140.00")
    assert dec(corps["reste_du"]) == Decimal("0.00")

    # 11 bis. §30 : la fiche client (vue-360) expose le reste dû réel PAR quittance, pas seulement
    # le TTC. La quittance réglée y remonte donc reste_du = 0 (et non 1140), montant_regle = 1140.
    reponse = client.get(f"/clients/{client_id}/vue-360", headers=auth)
    assert reponse.status_code == 200, reponse.text
    vue = reponse.json()
    q_vue = next((q for q in vue["quittances"] if q["id"] == quittance_id), None)
    assert q_vue is not None, vue  # la quittance doit figurer dans la vue client
    assert q_vue["reste_du"] is not None, q_vue  # enrichie, pas laissée à None
    assert dec(q_vue["reste_du"]) == Decimal("0.00"), q_vue
    assert dec(q_vue["montant_regle"]) == Decimal("1140.00"), q_vue

    # 12. Bordereau de versement, ajout de l'encaissement, validation (super admin)
    reponse = client.post("/banq/bordereaux", headers=auth, json={
        "banque_agence_id": refs["banque_id"], "date_bordereau": AUJ.isoformat(),
        "statut": "verse",  # tentative de contournement : ignorée, forcé brouillon côté serveur
    })
    assert reponse.status_code == 200, reponse.text
    assert reponse.json()["statut"] == "brouillon", reponse.json()
    bordereau_versement_id = reponse.json()["id"]
    reponse = client.post(f"/banq/bordereaux/{bordereau_versement_id}/ajouter/{encaissement_id}", headers=auth)
    assert reponse.status_code == 200, reponse.text
    reponse = client.patch(f"/banq/bordereaux/{bordereau_versement_id}/valider", headers=auth)
    assert reponse.status_code == 200, reponse.text
    assert reponse.json()["statut"] == "verse"

    # 13. L'encaissement passe en « rapproché banque » (RF-BANQ-02)
    reponse = client.get(f"/encaissements/{encaissement_id}", headers=auth)
    assert reponse.json()["statut"] == "rapproche_banque", reponse.json()

    # 14. Bordereau de reversement : sélection automatique de la période, puis validation
    reponse = client.post("/rev/bordereaux", headers=auth, json={
        "compagnie_id": refs["compagnie_id"],
        "periode_debut": (AUJ - datetime.timedelta(days=1)).isoformat(),
        "periode_fin": (AUJ + datetime.timedelta(days=1)).isoformat(),
        "statut": "reverse", "valide_par": 999,  # tentative de contournement : ignorée côté serveur
    })
    assert reponse.status_code == 200, reponse.text
    assert reponse.json()["statut"] == "brouillon", reponse.json()
    bordereau_reversement_id = reponse.json()["id"]

    reponse = client.post(f"/rev/bordereaux/{bordereau_reversement_id}/selectionner-periode", headers=auth)
    assert reponse.status_code == 200, reponse.text
    lignes = reponse.json()
    assert any(ligne["quittance_id"] == quittance_id for ligne in lignes), lignes

    reponse = client.patch(f"/rev/bordereaux/{bordereau_reversement_id}/valider", headers=auth)
    assert reponse.status_code == 200, reponse.text
    bordereau_reversement = reponse.json()

    # 15. Commission calculée selon le barème : 10 % de 1000 = 100 ; prime nette reversée = 1000
    assert dec(bordereau_reversement["commission_totale"]) == Decimal("100.00"), bordereau_reversement
    assert dec(bordereau_reversement["montant_total"]) == Decimal("1000.00"), bordereau_reversement

    # §28 : le GET de détail inclut les lignes, enrichies du numéro de quittance et de police
    reponse = client.get(f"/rev/bordereaux/{bordereau_reversement_id}", headers=auth)
    assert reponse.status_code == 200, reponse.text
    detail = reponse.json()
    assert len(detail["lignes"]) == 1, detail
    ligne_detail = detail["lignes"][0]
    assert ligne_detail["quittance_id"] == quittance_id
    assert ligne_detail["numero_quittance"] == quittance["numero_quittance"], ligne_detail
    assert ligne_detail["numero_police"] == police["numero_police"], ligne_detail

    # 15b. Reversement effectif (PATCH .../reverser, super admin, §18) : valide -> reverse
    url_reverser = f"/rev/bordereaux/{bordereau_reversement_id}/reverser"
    reponse = client.patch(url_reverser, headers=auth,  # référence vide -> 422
                           json={"reference_virement": "   ", "date_reversement": AUJ.isoformat()})
    assert reponse.status_code == 422, reponse.text
    reponse = client.patch(url_reverser, headers=auth,  # référence trop longue -> 422 (pas de 500)
                           json={"reference_virement": "X" * 101, "date_reversement": AUJ.isoformat()})
    assert reponse.status_code == 422, reponse.text
    reponse = client.patch(url_reverser, headers=auth,  # reversement effectif
                           json={"reference_virement": "VIR-2026-0001", "date_reversement": AUJ.isoformat()})
    assert reponse.status_code == 200, reponse.text
    reverse = reponse.json()
    assert reverse["statut"] == "reverse", reverse
    assert reverse["reference_virement"] == "VIR-2026-0001"
    assert reverse["date_reversement"] == AUJ.isoformat()
    reponse = client.patch(url_reverser, headers=auth,  # déjà reversé (plus « valide ») -> 400
                           json={"reference_virement": "VIR-2026-0002", "date_reversement": AUJ.isoformat()})
    assert reponse.status_code == 400, reponse.text

    # 16. Tableau de bord cohérent (base isolée : seules nos données existent)
    reponse = client.get("/dashboard", headers=auth)
    assert reponse.status_code == 200, reponse.text
    tableau = reponse.json()
    assert tableau["nombre_clients"] == 1, tableau
    assert tableau["contrats_actifs"] == 1, tableau               # police en vigueur (effet = aujourd'hui)
    assert tableau["paiements_en_attente"] == 0, tableau          # la seule quittance est réglée
    assert tableau["cheques_non_encaisses"] == 0, tableau         # le chèque est rapproché banque
    assert tableau["dossiers_recouvrement_ouverts"] == 0, tableau

    # 17. Éditions PDF (RF-POL-04) : génération côté serveur + archivage horodaté
    documents = [
        ("quittances", quittance_id),
        ("attestations", police_id),
        ("polices", police_id),
        ("bordereaux-versement", bordereau_versement_id),
        ("bordereaux-reversement", bordereau_reversement_id),
    ]
    for chemin, entite_id in documents:
        reponse = client.get(f"/documents/{chemin}/{entite_id}", headers=auth)
        assert reponse.status_code == 200, (chemin, reponse.text)
        assert reponse.headers["content-type"] == "application/pdf", (chemin, reponse.headers)
        assert reponse.content[:5] == b"%PDF-", chemin  # en-tête d'un PDF valide

    # Les 5 générations ont laissé une trace d'archive horodatée, re-téléchargeable
    reponse = client.get("/documents/archives", headers=auth)
    assert reponse.status_code == 200, reponse.text
    archives = reponse.json()
    assert len(archives) == 5, archives
    reponse = client.get(f"/documents/archives/{archives[0]['id']}/telecharger", headers=auth)
    assert reponse.status_code == 200 and reponse.content[:5] == b"%PDF-", reponse.status_code

    # 18. Profil de l'utilisateur connecté (GET /auth/me, écart §23)
    reponse = client.get("/auth/me", headers=auth)
    assert reponse.status_code == 200, reponse.text
    profil = reponse.json()
    assert profil["email"] == "admin@recette.test", profil
    assert profil["role"] == "super_administrateur", profil
    assert "mot_de_passe_hash" not in profil  # le hash ne fuit jamais

    # 19. Annulation de quittance (§11) + numéro de police dans la balance âgée (§14).
    # On monte une quittance dédiée, émise et non payée (après les comptes du dashboard).
    reponse = client.post("/clients", headers=auth, json={
        "type": "entreprise", "raison_sociale": "Annulation SARL", "ice": "ANNUL0001"})
    client_annul_id = reponse.json()["id"]
    reponse = client.post("/sous/polices", headers=auth, json={
        "client_id": client_annul_id, "produit_id": refs["produit_id"],
        "date_effet": AUJ.isoformat(), "date_echeance": DANS_UN_AN.isoformat()})
    police_annul = reponse.json()
    reponse = client.post("/risques", headers=auth, json={
        "police_id": police_annul["id"], "type_risque": "vehicule", "attributs": {
            "immatriculation": "99-B-99", "marque": "Fiat", "modele": "Panda",
            "date_mise_en_circulation": "01/2020", "valeur_neuf": "90000.00"}})
    reponse = client.post("/police-garanties", headers=auth, json={
        "police_id": police_annul["id"], "risque_id": reponse.json()["id"],
        "garantie_id": refs["garanties"]["DC"], "montant_prime": "500.00"})
    assert reponse.status_code == 200, reponse.text
    for piece_id in refs["pieces_oblig"]:
        client.post("/pieces-fournies", headers=auth, json={
            "police_id": police_annul["id"], "piece_requise_id": piece_id, "reference": "ok"})
    reponse = client.post("/sous/avenants", headers=auth, json={
        "police_id": police_annul["id"], "type_avenant": "affaire_nouvelle", "date_effet": AUJ.isoformat()})
    reponse = client.patch(f"/sous/avenants/{reponse.json()['id']}/valider", headers=auth)
    quittance_annul_id = reponse.json()["quittance"]["id"]

    # §14 : la quittance impayée figure dans la balance âgée, avec son numéro de police
    reponse = client.get("/recouv/balance-agee", headers=auth)
    assert reponse.status_code == 200, reponse.text
    ligne = next((l for l in reponse.json()["lignes"] if l["quittance_id"] == quittance_annul_id), None)
    assert ligne is not None and ligne["numero_police"] == police_annul["numero_police"], ligne

    # §11 : annulation réservée super admin, motif obligatoire, uniquement si émise
    reponse = client.patch(f"/quittances/{quittance_annul_id}/annuler", headers=auth,
                           json={"motif": "   "})  # motif vide -> 422
    assert reponse.status_code == 422, reponse.text
    reponse = client.patch(f"/quittances/{quittance_annul_id}/annuler", headers=auth,
                           json={"motif": "Erreur de saisie"})
    assert reponse.status_code == 200, reponse.text
    annulee = reponse.json()
    assert annulee["statut"] == "annulee" and annulee["motif_annulation"] == "Erreur de saisie", annulee
    assert dec(annulee["reste_du"]) == Decimal("0.00"), annulee  # annulée -> plus rien dû
    reponse = client.patch(f"/quittances/{quittance_annul_id}/annuler", headers=auth,
                           json={"motif": "déjà annulée"})  # plus émise -> 400
    assert reponse.status_code == 400, reponse.text

    # Le chiffre d'affaires exclut la quittance annulée (seule la quittance réglée de 1140 compte)
    reponse = client.get("/reporting/chiffre-affaires", headers=auth, params={
        "date_debut": (AUJ - datetime.timedelta(days=1)).isoformat(),
        "date_fin": (AUJ + datetime.timedelta(days=1)).isoformat()})
    assert reponse.status_code == 200, reponse.text
    ca = reponse.json()
    assert ca["nombre_quittances"] == 1, ca
    assert dec(ca["prime_ttc_totale"]) == Decimal("1140.00"), ca

    # 21. Avenant de modification (règle 17) : sémantique minimale et tracée, sans quittance.
    # La police principale a un avenant « affaire nouvelle » validé -> composition verrouillée :
    # retirer directement une garantie est refusé (règle 13).
    reponse = client.get("/police-garanties", headers=auth, params={"police_id": police_id})
    pg_id = reponse.json()[-1]["id"]  # une garantie ajoutée en 8b, hors quittance déjà émise
    reponse = client.delete(f"/police-garanties/{pg_id}", headers=auth)
    assert reponse.status_code == 400, reponse.text  # verrouillée : passer par un avenant

    # Un avenant de modification sans motif est refusé (422) : le motif est la seule trace du
    # « quoi » modifié tant qu'il n'y a pas de champs structurés.
    reponse = client.post("/sous/avenants", headers=auth, json={
        "police_id": police_id, "type_avenant": "modification", "date_effet": AUJ.isoformat()})
    assert reponse.status_code == 422, reponse.text

    reponse = client.post("/sous/avenants", headers=auth, json={
        "police_id": police_id, "type_avenant": "modification", "date_effet": AUJ.isoformat(),
        "motif": "Retrait d'une garantie"})
    assert reponse.status_code == 200, reponse.text
    avenant_modif_id = reponse.json()["id"]
    assert reponse.json()["statut"] == "brouillon", reponse.json()

    # Brouillon de modification en cours -> la composition est rouverte : le retrait passe.
    reponse = client.delete(f"/police-garanties/{pg_id}", headers=auth)
    assert reponse.status_code == 204, reponse.text

    # Validation de la modification : aucune quittance générée (volet financier différé),
    # et la composition se reverrouille.
    reponse = client.patch(f"/sous/avenants/{avenant_modif_id}/valider", headers=auth)
    assert reponse.status_code == 200, reponse.text
    assert reponse.json()["quittance"] is None, reponse.json()

    reponse = client.get("/police-garanties", headers=auth, params={"police_id": police_id})
    autre_pg_id = reponse.json()[-1]["id"]
    reponse = client.delete(f"/police-garanties/{autre_pg_id}", headers=auth)
    assert reponse.status_code == 400, reponse.text  # reverrouillée après validation

    # 22. Exports Excel des rapports (section 5.3) : mêmes données qu'en JSON, en .xlsx serveur.
    periode = {"date_debut": (AUJ - datetime.timedelta(days=1)).isoformat(),
               "date_fin": (AUJ + datetime.timedelta(days=1)).isoformat()}

    # format=json (défaut) inchangé : le reporting reste consultable en JSON
    reponse = client.get("/reporting/chiffre-affaires", headers=auth, params=periode)
    assert reponse.status_code == 200 and reponse.headers["content-type"].startswith("application/json")

    # Chiffre d'affaires en xlsx : valeurs scalaires dans la feuille « Résumé »
    reponse = client.get("/reporting/chiffre-affaires", headers=auth,
                         params={**periode, "format": "xlsx"})
    assert reponse.status_code == 200, reponse.text
    assert reponse.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert reponse.content[:2] == b"PK", "un .xlsx est une archive ZIP (signature PK)"
    classeur = load_workbook(BytesIO(reponse.content))
    resume = classeur["Résumé"]
    valeurs = {r[0].value: r[1].value for r in resume.iter_rows(min_row=4) if r[0].value}
    assert valeurs["Nombre quittances"] == 1, valeurs
    assert valeurs["Prime ttc totale"] == 1140.0, valeurs

    # Paiements en xlsx : le sous-dictionnaire par_mode_paiement est aplati en lignes préfixées
    reponse = client.get("/reporting/paiements", headers=auth, params={**periode, "format": "xlsx"})
    assert reponse.status_code == 200, reponse.text
    resume = load_workbook(BytesIO(reponse.content))["Résumé"]
    valeurs = {r[0].value: r[1].value for r in resume.iter_rows(min_row=4) if r[0].value}
    assert valeurs["Nombre encaissements"] == 1, valeurs
    assert any(str(cle).startswith("Par mode paiement - ") for cle in valeurs), valeurs

    # Échéances en xlsx : la liste des polices devient une feuille dédiée
    reponse = client.get("/reporting/echeances", headers=auth, params={"jours": 400, "format": "xlsx"})
    assert reponse.status_code == 200, reponse.text
    classeur = load_workbook(BytesIO(reponse.content))
    assert "Polices" in classeur.sheetnames, classeur.sheetnames
    assert classeur["Polices"].max_row >= 2, "en-tête + au moins une police"

    # Liste VIDE : la feuille dédiée existe quand même (cohérence avec le JSON qui garde la clé).
    # jours=0 -> horizon réduit à aujourd'hui, aucune police (échéances à un an) -> polices vides.
    reponse = client.get("/reporting/echeances", headers=auth, params={"jours": 0, "format": "xlsx"})
    assert reponse.status_code == 200, reponse.text
    classeur = load_workbook(BytesIO(reponse.content))
    assert "Polices" in classeur.sheetnames, classeur.sheetnames  # présente même à vide

    # Auth : les endpoints de reporting exigent un jeton (données financières, non publiques)
    reponse = client.get("/reporting/production", params=periode)  # sans en-tête d'autorisation
    assert reponse.status_code == 401, reponse.text
