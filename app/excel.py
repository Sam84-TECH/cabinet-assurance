"""
Export Excel des rapports de reporting (section 5.3 / DIAM Module 13).

Convertit le dict d'un rapport (même structure que la réponse JSON de `reporting.py`) en un
classeur .xlsx généré côté serveur. openpyxl : pur Python, sans dépendance système (comme
fpdf2 pour les PDF), conforme à la contrainte « serveur modeste » du CDCF.

Générique : une feuille « Résumé » pour les valeurs scalaires (et les sous-dictionnaires,
comme la répartition par mode de paiement), plus une feuille dédiée par liste de lignes
(ex. les polices arrivant à échéance). Aucun rapport n'a de code spécifique ici — `reporting.py`
lui passe simplement ses données, dans les types Python bruts (Decimal, date), avant toute
sérialisation JSON.
"""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

NAVY = "12253E"  # #12253E — couleur principale de la charte
_ENTETE_FILL = PatternFill("solid", fgColor=NAVY)
_ENTETE_FONT = Font(bold=True, color="FFFFFF")
_TITRE_FONT = Font(bold=True, color=NAVY, size=14)
_FORMAT_NOMBRE = "#,##0.00"
_FORMAT_DATE = "DD/MM/YYYY"
_CARACTERES_INTERDITS = set("[]:*?/\\")


def _humaniser(cle) -> str:
    """`nombre_polices` -> `Nombre polices` : libellé lisible à partir d'une clé technique."""
    return str(cle).replace("_", " ").strip().capitalize()


def _scalaire(valeur) -> bool:
    return valeur is None or isinstance(valeur, (str, bool, int, float, Decimal, date, datetime))


def _cellule(valeur):
    """Valeur prête pour openpyxl : Decimal -> float (openpyxl n'écrit pas les Decimal),
    None -> chaîne vide ; date/datetime/int/str sont gérés nativement."""
    if valeur is None:
        return ""
    if isinstance(valeur, Decimal):
        return float(valeur)
    return valeur


def _formater(cellule, valeur) -> None:
    # bool est un sous-type de int : on l'exclut du format nombre.
    if isinstance(valeur, (float, Decimal)) or (isinstance(valeur, int) and not isinstance(valeur, bool)):
        cellule.number_format = _FORMAT_NOMBRE
    elif isinstance(valeur, (date, datetime)):
        cellule.number_format = _FORMAT_DATE


def _kv(ws, label: str, valeur) -> None:
    """Ajoute une ligne « libellé / valeur » et formate la cellule de valeur."""
    ws.append([label, _cellule(valeur)])
    _formater(ws.cell(row=ws.max_row, column=2), valeur)


def _styler_entete(ws, ligne: int, nb_colonnes: int) -> None:
    for col in range(1, nb_colonnes + 1):
        cellule = ws.cell(row=ligne, column=col)
        cellule.fill = _ENTETE_FILL
        cellule.font = _ENTETE_FONT


def _ajuster_largeurs(ws) -> None:
    for colonne in ws.columns:
        longueur = max((len(str(c.value)) for c in colonne if c.value not in (None, "")), default=10)
        ws.column_dimensions[get_column_letter(colonne[0].column)].width = min(longueur + 4, 50)


def _nom_feuille(cle: str, pris: set[str]) -> str:
    """Nom de feuille valide et unique : 31 caractères max, sans caractères interdits, avec
    un suffixe de dédoublonnage (`_2`, `_3`…) si le nom est déjà pris (openpyxl refuse les
    doublons). `pris` accumule les noms déjà attribués."""
    propre = "".join(c for c in _humaniser(cle) if c not in _CARACTERES_INTERDITS)[:31] or "Détail"
    nom = propre
    suffixe = 2
    while nom in pris:
        rallonge = f"_{suffixe}"
        nom = propre[:31 - len(rallonge)] + rallonge
        suffixe += 1
    pris.add(nom)
    return nom


def _feuille_liste(wb: Workbook, cle: str, lignes: list[dict], pris: set[str]) -> None:
    """Écrit une liste de lignes dans une feuille dédiée. Une liste vide crée quand même la
    feuille (en-tête seul si les colonnes sont connues, sinon une note) : le xlsx garde la
    même structure que le JSON, qui conserve la clé même à vide (point de cohérence)."""
    ws = wb.create_sheet(title=_nom_feuille(cle, pris))
    if not lignes:
        ws.append(["(aucune ligne)"])
        _ajuster_largeurs(ws)
        return
    colonnes = list(lignes[0].keys())
    ws.append([_humaniser(c) for c in colonnes])
    _styler_entete(ws, 1, len(colonnes))
    for ligne in lignes:
        valeurs = [ligne.get(c) for c in colonnes]
        ws.append([_cellule(v) for v in valeurs])
        for idx, valeur in enumerate(valeurs, start=1):
            _formater(ws.cell(row=ws.max_row, column=idx), valeur)
    _ajuster_largeurs(ws)


def rapport_vers_xlsx(titre: str, donnees: dict) -> bytes:
    """Rend `donnees` (dict d'un rapport) en classeur .xlsx et renvoie les octets du fichier."""
    wb = Workbook()
    resume = wb.active
    resume.title = "Résumé"

    resume.append([titre])
    resume["A1"].font = _TITRE_FONT
    resume.append([])  # ligne d'aération
    resume.append(["Indicateur", "Valeur"])
    _styler_entete(resume, resume.max_row, 2)

    listes: dict[str, list[dict]] = {}
    for cle, valeur in donnees.items():
        if _scalaire(valeur):
            _kv(resume, _humaniser(cle), valeur)
        elif isinstance(valeur, dict):
            # sous-dictionnaire (ex. par_mode_paiement) : une ligne par entrée, préfixée.
            for sous_cle, sous_val in valeur.items():
                _kv(resume, f"{_humaniser(cle)} - {_humaniser(sous_cle)}", sous_val)
        elif isinstance(valeur, list) and all(isinstance(x, dict) for x in valeur):
            # liste de lignes (même VIDE : `all(...)` est vrai sur une liste vide) -> feuille
            # dédiée, pour garder la même structure que le JSON qui conserve la clé à vide.
            listes[cle] = valeur
        elif isinstance(valeur, list):
            _kv(resume, _humaniser(cle), ", ".join(str(x) for x in valeur))

    pris_feuilles = {resume.title}  # « Résumé » est déjà pris : les feuilles listes s'en écartent
    for cle, lignes in listes.items():
        _feuille_liste(wb, cle, lignes, pris_feuilles)

    _ajuster_largeurs(resume)
    tampon = BytesIO()
    wb.save(tampon)
    return tampon.getvalue()
