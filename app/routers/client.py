"""
Module Client (CRM minimal) — une seule table, particulier + entreprise (option A validée).
"""

import unicodedata
from decimal import Decimal
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy.exc import DataError, IntegrityError
from sqlalchemy.orm import Session

from .. import crud, models, schemas
from ..database import get_db
from ..models import TypeClient
from ..reglement import enrichir_quittance
from ..auth import get_current_user

router = APIRouter(prefix="/clients", tags=["Client"])


# ------------------------------------------------------------
# Import Excel — normalisation des en-têtes et du type de client
# ------------------------------------------------------------

def _normaliser(valeur) -> str:
    """Minuscule, sans accents ni espaces superflus (« Raison Sociale » -> « raison_sociale »).
    Sert à reconnaître les en-têtes de colonnes et les valeurs de type quel que soit leur casse."""
    if valeur is None:
        return ""
    sans_accent = unicodedata.normalize("NFKD", str(valeur)).encode("ascii", "ignore").decode()
    return sans_accent.strip().lower().replace(" ", "_")


# En-tête normalisé de colonne -> champ client. Plusieurs libellés acceptés pour un même champ.
ENTETES_CLIENT = {
    "nom": "nom",
    "prenom": "prenom",
    "raison_sociale": "raison_sociale",
    "raisonsociale": "raison_sociale",
    "cin": "cin",
    "ice": "ice",
    "telephone": "telephone",
    "tel": "telephone",
    "ville": "ville",
    "type": "type",
    "email": "email",
    "adresse": "adresse",
}


def _type_client(brut) -> TypeClient | None:
    """Déduit le type d'un libellé libre : « entreprise/PM/société/morale » -> entreprise,
    « particulier/PP/physique » -> particulier. None si non reconnu (l'appelant infère alors)."""
    n = _normaliser(brut)
    if not n:
        return None
    if n == "pm" or n[0] == "e" or "moral" in n or "societ" in n:
        return TypeClient.entreprise
    if n[0] == "p" or "physiq" in n:
        return TypeClient.particulier
    return None


@router.post("", response_model=schemas.ClientRead)
def create_client(payload: schemas.ClientCreate, db: Session = Depends(get_db),
                   user: models.Utilisateur = Depends(get_current_user)):
    return crud.create(db, models.Client, payload.model_dump())


@router.post("/import-excel", response_model=schemas.ResultatImportClients)
def importer_clients_excel(fichier: UploadFile = File(...), db: Session = Depends(get_db),
                           user: models.Utilisateur = Depends(get_current_user)):
    """Création en masse de clients depuis un classeur Excel (.xlsx). La première ligne est
    l'en-tête ; colonnes reconnues (casse/accents indifférents) : nom, prenom, raison_sociale,
    cin, ice, telephone, ville, type, email, adresse. Chaque ligne est validée comme un client
    normal (mêmes règles d'identité : particulier -> CIN + nom + prénom ; entreprise -> raison
    sociale + ICE) puis insérée dans un point de sauvegarde isolé : une ligne fautive (doublon
    CIN/ICE, champ obligatoire manquant, type invalide) est rejetée et rapportée SANS annuler les
    autres. Renvoie un rapport (lignes lues, créées, erreurs ligne par ligne)."""
    # Plafonds : l'app tourne en agence sur un serveur modeste ; on borne la taille et le nombre
    # de lignes pour ne pas saturer la RAM avec un fichier envoyé par un utilisateur authentifié.
    TAILLE_MAX = 2 * 1024 * 1024  # 2 Mo
    LIGNES_MAX = 5000
    contenu = fichier.file.read()
    if len(contenu) > TAILLE_MAX:
        raise HTTPException(400, "Fichier trop volumineux (maximum 2 Mo).")

    try:
        classeur = load_workbook(BytesIO(contenu), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "Fichier illisible : un classeur Excel (.xlsx) est attendu.")

    feuille = classeur.active
    if feuille is None:
        raise HTTPException(400, "Le classeur ne contient aucune feuille exploitable.")
    lignes = list(feuille.iter_rows(values_only=True))
    if not lignes:
        raise HTTPException(400, "Le fichier est vide.")
    if len(lignes) - 1 > LIGNES_MAX:
        raise HTTPException(400, f"Trop de lignes (maximum {LIGNES_MAX}). Découpez le fichier en plusieurs imports.")

    # En-tête -> index de colonne (on ne garde que les colonnes reconnues).
    colonnes = {}
    for index, brut in enumerate(lignes[0]):
        champ = ENTETES_CLIENT.get(_normaliser(brut))
        if champ is not None:
            colonnes[champ] = index
    if not colonnes:
        raise HTTPException(
            400, "Aucune colonne reconnue en en-tête. Attendu : nom, prenom, raison_sociale, "
                 "cin, ice, telephone, ville, type.")

    def cellule(ligne, champ):
        index = colonnes.get(champ)
        if index is None or index >= len(ligne) or ligne[index] is None:
            return ""
        return str(ligne[index]).strip()

    erreurs: list[dict] = []
    noms_crees: list[str] = []
    lignes_donnees = 0

    for numero, ligne in enumerate(lignes[1:], start=2):  # ligne 2 = première ligne de données
        if all(c is None or str(c).strip() == "" for c in ligne):
            continue  # ligne entièrement vide : ignorée sans erreur
        lignes_donnees += 1

        libelle = cellule(ligne, "raison_sociale") or " ".join(
            p for p in (cellule(ligne, "nom"), cellule(ligne, "prenom")) if p) or None

        # Type : depuis la colonne, sinon inféré (raison sociale/ICE -> entreprise, sinon particulier).
        type_client = _type_client(cellule(ligne, "type"))
        if type_client is None:
            type_client = (
                TypeClient.entreprise
                if (cellule(ligne, "raison_sociale") or cellule(ligne, "ice"))
                else TypeClient.particulier
            )

        donnees = {"type": type_client}
        for champ in ("telephone", "ville", "email", "adresse", "nom", "prenom", "cin",
                      "raison_sociale", "ice"):
            valeur = cellule(ligne, champ)
            if valeur:
                donnees[champ] = valeur

        # Validation identique à une création unitaire (règles d'identité selon le type).
        try:
            payload = schemas.ClientCreate(**donnees)
        except ValidationError as err:
            message = err.errors()[0].get("msg", "Données invalides.")
            message = message.replace("Value error, ", "")  # préfixe technique de Pydantic
            erreurs.append({"ligne": numero, "valeur": libelle, "message": message})
            continue

        # Insertion isolée par point de sauvegarde : un doublon CIN/ICE ne casse pas le lot.
        try:
            with db.begin_nested():
                client = models.Client(**payload.model_dump())
                db.add(client)
                db.flush()
            noms_crees.append(libelle or f"Client #{client.id}")
        except IntegrityError:
            champ_doublon = "ICE" if type_client == TypeClient.entreprise else "CIN"
            erreurs.append({
                "ligne": numero, "valeur": libelle,
                "message": f"Doublon : ce {champ_doublon} existe déjà (client non créé).",
            })
        except DataError:
            # Valeur trop longue / mal typée pour une colonne (ex. CIN > 20 caractères) : le
            # point de sauvegarde est annulé, la ligne rejetée et le lot continue — sinon un 500
            # non rattrapé ferait perdre tout l'import (get_db rollback en fin de requête).
            erreurs.append({
                "ligne": numero, "valeur": libelle,
                "message": "Valeur trop longue ou invalide pour un champ (vérifiez longueurs et formats).",
            })

    db.commit()
    return {
        "lignes_totales": lignes_donnees,
        "crees": len(noms_crees),
        "noms_crees": noms_crees,
        "erreurs": erreurs,
    }


@router.get("", response_model=list[schemas.ClientRead])
def list_clients(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    return crud.list_all(db, models.Client, skip, limit)


@router.get("/{client_id}", response_model=schemas.ClientRead)
def get_client(client_id: int, db: Session = Depends(get_db)):
    return crud.get_or_404(db, models.Client, client_id)


@router.patch("/{client_id}", response_model=schemas.ClientRead)
def update_client(client_id: int, payload: schemas.ClientUpdate, db: Session = Depends(get_db),
                   user: models.Utilisateur = Depends(get_current_user)):
    return crud.update(db, models.Client, client_id, payload.model_dump(exclude_unset=True))


@router.delete("/{client_id}", status_code=204)
def delete_client(client_id: int, db: Session = Depends(get_db),
                   user: models.Utilisateur = Depends(get_current_user)):
    """Suppression refusée si le client est référencé ailleurs — police, encaissement,
    lien familial… — via le contrôle générique de `crud.delete` (règles 14 et 15)."""
    crud.delete(db, models.Client, client_id)


@router.get("/{client_id}/vue-360", response_model=schemas.Vue360Client)
def vue_360(client_id: int, db: Session = Depends(get_db),
            user: models.Utilisateur = Depends(get_current_user)):
    """Vue 360 du client (RF-CRM-06, RF-ENC-05) : ses polices, ses quittances, son solde
    (total dû sur quittances non annulées, total encaissé hors chèques rejetés, reste dû)
    et l'historique de ses encaissements."""
    client = crud.get_or_404(db, models.Client, client_id)
    polices = db.query(models.Police).filter_by(client_id=client_id).all()
    quittances = (
        db.query(models.Quittance).join(models.Police)
        .filter(models.Police.client_id == client_id).all()
    )
    # §30 (même cause que §27) : chaque quittance de la fiche client expose son reste dû réel
    # (prime_ttc - réglé hors chèques rejetés), pas seulement son TTC total.
    for q in quittances:
        enrichir_quittance(db, q)
    encaissements = db.query(models.Encaissement).filter_by(client_id=client_id).all()

    total_du = sum(
        (q.prime_ttc for q in quittances if q.statut != models.StatutQuittance.annulee),
        Decimal("0"),
    )
    total_encaisse = sum(
        (e.montant for e in encaissements if e.statut != models.StatutEncaissement.rejete),
        Decimal("0"),
    )
    return {
        "client": client,
        "polices": polices,
        "quittances": quittances,
        "solde": {
            "total_du": total_du,
            "total_encaisse": total_encaisse,
            "reste_du": total_du - total_encaisse,
        },
        "encaissements": encaissements,
    }
